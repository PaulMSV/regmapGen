
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ind 
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


def parse_bits(bits):
    try:
        start = int(bits)
        return [(start, start)]
    except ValueError:
        bit_n = [int(x) for x in bits.split(':')]
        start = min(bit_n)
        end = max(bit_n)
        return [(i, i) for i in range(start, end + 1)]


def calculate_address(block_offset, reg_offset):
    if block_offset is not None and reg_offset is not None:
        try:
            return hex(int(block_offset, 16) + int(reg_offset, 16))
        except ValueError:
            return None
    return None


def flat_xls_rmap(input_file, output_file, include_reset, include_desc):
    # Load the uploaded Excel file using openpyxl
    wb = load_workbook(input_file, data_only=True)
    ws = wb['RegMap']

    # Initialize a dictionary to collect register data
    reg_dict = {}
    access_dict = {}
    desc_dict = {}

    current_reg = None
    current_block_offset = None

    # Locate columns explicitly
    block_offset_col = 'C'
    reg_offset_col = 'F'
    access_col = 'G'
    reg_col = 'E'
    field_col = 'J'
    bits_col = 'K'
    reset_col = 'M'
    desc_col = 'I'

    for row in ws.iter_rows(min_row=2, values_only=True):
        if pd.notna(row[ind('A') - 1]):
            current_block_offset = row[ind(block_offset_col) - 1]
        if pd.notna(row[ind(reg_col) - 1]):
            current_reg = row[ind(reg_col) - 1]
            reg_offset = row[ind(reg_offset_col) - 1]
            address = calculate_address(current_block_offset, reg_offset)
            access = row[ind(access_col) - 1]
            desc = row[ind(desc_col) - 1] if include_desc else ""
            reg_dict[current_reg] = ["reserved"] * 8  # Initialize with "reserved"
            access_dict[current_reg] = {"Address": address, "Access": access}
            desc_dict[current_reg] = desc
        elif current_reg and pd.notna(row[ind(field_col) - 1]) and pd.notna(row[ind(bits_col) - 1]):
            field = row[ind(field_col) - 1]
            bits = row[ind(bits_col) - 1]
            reset_value = row[ind(reset_col) - 1]
            if include_reset and reset_value is not None:
                field_with_reset = f"{field} ({reset_value})"
            else:
                field_with_reset = field
            
            # Parse the bit positions and assign the field names
            for bit_range in bits.split(','):
                bit_range = bit_range.strip()
                parsed_bits = parse_bits(bit_range)
                for bit in parsed_bits:
                    reg_dict[current_reg][7 - bit[0]] = field_with_reset  # Bits are in reverse order for columns

    # Create a DataFrame from the collected data
    transformed_data = []
    for reg, bits in reg_dict.items():
        address = access_dict[reg]["Address"]
        access = access_dict[reg]["Access"]
        if include_desc:
            desc = desc_dict[reg]
            transformed_data.append([address, access, reg] + bits + [desc])
        else:
            transformed_data.append([address, access, reg] + bits)
    
    if include_desc:
        columns = ['Address', 'Access', 'Register', 'Bit[7]', 'Bit[6]', 'Bit[5]', 'Bit[4]', 'Bit[3]', 'Bit[2]', 'Bit[1]', 'Bit[0]', 'Description']
    else:
        columns = ['Address', 'Access', 'Register', 'Bit[7]', 'Bit[6]', 'Bit[5]', 'Bit[4]', 'Bit[3]', 'Bit[2]', 'Bit[1]', 'Bit[0]']
    
    transformed_df = pd.DataFrame(transformed_data, columns=columns)

    # Save the transformed dataframe to an Excel file
    transformed_df.to_excel(output_file, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Set the width of specific columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    if include_desc:
        ws.column_dimensions['L'].width = 30  # Set width for Description column
    for col in ws.columns:
        if col[0].column_letter not in ['A', 'B', 'C'] + (['L'] if include_desc else []):
            ws.column_dimensions[col[0].column_letter].width = 25

    # Apply fill to header
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Apply fill and bold font to register names
    reg_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    reg_font = Font(bold=True)
    for row in range(2, len(transformed_df) + 2):
        cell = ws.cell(row=row, column=3)  # Column 'Register'
        cell.fill = reg_fill
        cell.font = reg_font

    # Center align all cells in the Bit 7 to Bit 0 columns and apply border to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(2, len(transformed_df) + 2):
        for col in range(4, 12):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if cell.value == "not_used":
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Apply border to all other cells
    max_col = 12 if include_desc else 11
    for row in ws.iter_rows(min_row=1, max_row=len(transformed_df) + 1, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    # Iterate over the dataframe and merge cells where necessary
    for row in range(2, len(transformed_df) + 2):
        for col in range(4, 12):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str) and cell_value != "not_used":
                # Determine the range of merged cells
                start_col = col
                end_col = col
                while end_col < 12 and ws.cell(row=row, column=end_col + 1).value == cell_value:
                    end_col += 1
                
                if start_col != end_col:
                    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                    ws.cell(row=row, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
                    col = end_col  # Move to the next unmerged cell

    # Save the changes
    wb.save(output_file)

    print(f"Transformation complete. The transformed file is saved as {output_file}.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Transform register map Excel file.')
    parser.add_argument('-i', '--input', required=True, help='Input Excel file')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser.add_argument('-rst', '--include_reset', action='store_true', help='Include reset values in bit field names')
    parser.add_argument('-desc', '--include_desc', action='store_true', help='Include descriptions in the output file')
    
    args = parser.parse_args()
    
    flat_xls_rmap(args.input, args.output, args.include_reset, args.include_desc)
